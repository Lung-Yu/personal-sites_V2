#!/usr/bin/env python3
"""Read and analyze 年度績效評核表 xlsx files."""

import glob
import openpyxl
from pathlib import Path
import sys


def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {"file": Path(path).name, "sheets": {}}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(list(row))
        result["sheets"][sheet_name] = rows
    return result


def analyze_all(folder="."):
    pattern = folder + "/**/*.xlsx"
    files = glob.glob(pattern, recursive=True)
    files = [f for f in files if "\u7e3e\u6548\u8a55\u6838" in f]

    if not files:
        print("No xlsx files found in: " + folder)
        return

    for path in sorted(files):
        print("\n" + "=" * 70)
        print("FILE: " + Path(path).name)
        print("=" * 70)
        data = read_xlsx(path)
        for sheet_name, rows in data["sheets"].items():
            print("\n  [Sheet: " + sheet_name + "]")
            for row in rows:
                cells = [str(c) for c in row if c is not None]
                if cells:
                    print("    " + " | ".join(cells))


folder = sys.argv[1] if len(sys.argv) > 1 else "."
analyze_all(folder)
